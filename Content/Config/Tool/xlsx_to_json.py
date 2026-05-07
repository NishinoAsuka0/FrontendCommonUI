"""
xlsx_to_json.py

将 Content/Config/Data/ 下的 .xlsx 文件转换为 UE5 DataTable 兼容的 JSON，
输出到 Content/Config/Json/。

列名映射：
  - 第一列 "Name" → JSON "Name" (DataTable RowName)
  - (Tag=xxx) 格式 → 提取 "xxx"
  - (Tags=xxx|yyy) 格式 → 提取 "xxx|yyy"
  - NSLOCTEXT(,,text) → 保留原样，UE 自动解析
  - 数值/布尔 → 按原类型写入 JSON

用法：
  python xlsx_to_json.py
"""

import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    print("需要 openpyxl 库: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(SCRIPT_DIR)
DATA_DIR = os.path.join(ROOT_DIR, "Data")
JSON_DIR = os.path.join(ROOT_DIR, "Json")


def parse_number(s: str):
    """尝试将字符串解析为 int 或 float"""
    s = s.strip()
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        pass
    return None


def parse_cell_value(value):
    """解析单元格值，识别 UE 特有格式"""
    if value is None:
        return None

    # 数值类型直接返回
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value == int(value):
            return int(value)
        return value

    # 布尔类型
    if isinstance(value, bool):
        return value

    s = str(value).strip()
    if not s:
        return None

    # 布尔字符串
    if s.lower() == "true":
        return True
    if s.lower() == "false":
        return False

    # (Tag=xxx) → 提取 Tag 字符串
    m = re.match(r"^\(Tag=(.+)\)$", s)
    if m:
        return m.group(1)

    # (Tags=xxx|yyy) → 提取 Tags 字符串
    m = re.match(r"^\(Tags=(.+)\)$", s)
    if m:
        return m.group(1)

    # 纯数字字符串
    num = parse_number(s)
    if num is not None:
        return num

    # NSLOCTEXT 和其他文本保持原样
    return s


def convert_xlsx(xlsx_path: str, json_path: str) -> int:
    """转换单个 xlsx → json，返回行数"""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    if len(rows) < 2:
        print(f"  跳过: 无数据行")
        return 0

    headers = [str(c) if c is not None else "" for c in rows[0]]
    # 去除可能的空白列尾
    while headers and headers[-1] == "":
        headers.pop()

    result = []
    for row_idx, row in enumerate(rows[1:], start=2):
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        obj = {}
        for col_idx in range(len(headers)):
            value = row[col_idx] if col_idx < len(row) else None
            header = headers[col_idx]

            if header == "Name":
                # RowName: 保证是字符串
                obj["Name"] = str(int(value)) if isinstance(value, float) else str(value)
            else:
                parsed = parse_cell_value(value)
                if parsed is not None:
                    obj[header] = parsed

        if obj.get("Name"):
            result.append(obj)
        else:
            print(f"  警告: 第 {row_idx} 行 Name 为空，已跳过")

    os.makedirs(os.path.dirname(json_path), exist_ok=True)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    return len(result)


def main():
    if not os.path.isdir(DATA_DIR):
        print(f"错误: Data 目录不存在: {DATA_DIR}")
        sys.exit(1)

    os.makedirs(JSON_DIR, exist_ok=True)

    total = 0
    for name in sorted(os.listdir(DATA_DIR)):
        if not name.endswith(".xlsx"):
            continue

        xlsx_path = os.path.join(DATA_DIR, name)
        json_name = name.replace(".xlsx", ".json")
        json_path = os.path.join(JSON_DIR, json_name)

        count = convert_xlsx(xlsx_path, json_path)
        print(f"  {name} → {json_name}  ({count} 行)")
        total += count

    print(f"\n完成: 共生成 {total} 行数据")


if __name__ == "__main__":
    main()
